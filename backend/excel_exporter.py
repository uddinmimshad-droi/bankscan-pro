from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from parser import Transaction
from utils import date_to_display


HEADERS = ["TRANSACTION DATE", "PARTICULARS", "WITHDRAW", "DEPOSIT", "BALANCE"]


def export_transactions_to_excel(
    transactions: List[Transaction],
    output_path: Path,
    pages_processed: int,
    ocr_pages: int,
    digital_pages: int,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    _write_transactions_sheet(sheet, transactions)
    _write_summary_sheet(workbook, transactions, pages_processed, ocr_pages, digital_pages)
    workbook.save(output_path)
    return output_path


def _write_transactions_sheet(sheet, transactions: List[Transaction]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    grey_fill = PatternFill("solid", fgColor="F3F6FA")
    red_font = Font(color="C00000")
    green_font = Font(color="008000")

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, tx in enumerate(transactions, start=2):
        sheet.append([tx.transaction_date, tx.particulars, tx.withdraw, tx.deposit, tx.balance])
        fill = grey_fill if row_index % 2 == 0 else white_fill
        for cell in sheet[row_index]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top")
        sheet.cell(row_index, 1).number_format = "DD/MM/YYYY"
        sheet.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        for column in (3, 4, 5):
            sheet.cell(row_index, column).number_format = '#,##0.00'
        if tx.withdraw and tx.withdraw > 0:
            sheet.cell(row_index, 3).font = red_font
        if tx.deposit and tx.deposit > 0:
            sheet.cell(row_index, 4).font = green_font

    summary_row = len(transactions) + 3
    sheet.cell(summary_row, 2, "Totals")
    sheet.cell(summary_row, 3, f"=SUM(C2:C{len(transactions) + 1})")
    sheet.cell(summary_row, 4, f"=SUM(D2:D{len(transactions) + 1})")
    sheet.cell(summary_row, 5, transactions[-1].balance if transactions else 0)
    for column in range(2, 6):
        cell = sheet.cell(summary_row, column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        if column >= 3:
            cell.number_format = '#,##0.00'

    sheet.freeze_panes = "A2"
    _auto_fit(sheet)


def _write_summary_sheet(workbook: Workbook, transactions: List[Transaction], pages_processed: int, ocr_pages: int, digital_pages: int) -> None:
    sheet = workbook.create_sheet("Summary")
    total_debits = sum((tx.withdraw or 0 for tx in transactions))
    total_credits = sum((tx.deposit or 0 for tx in transactions))
    first_date = date_to_display(transactions[0].transaction_date) if transactions else ""
    last_date = date_to_display(transactions[-1].transaction_date) if transactions else ""
    rows = [
        ("Total number of transactions", len(transactions)),
        ("Date range", f"{first_date} to {last_date}" if transactions else ""),
        ("Total Credits", total_credits),
        ("Total Debits", total_debits),
        ("Net Flow", total_credits - total_debits),
        ("Number of pages processed", pages_processed),
        ("Pages using OCR", ocr_pages),
        ("Pages using digital text", digital_pages),
    ]
    sheet.append(["Metric", "Value"])
    for metric, value in rows:
        sheet.append([metric, value])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row[0].font = Font(bold=True)
    for row in range(4, 7):
        sheet.cell(row, 2).number_format = '#,##0.00'
    _auto_fit(sheet)


def _auto_fit(sheet) -> None:
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = min(max(length + 2, 12), 60)
